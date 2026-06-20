"""XLSX backend.

A cell whose text matches the question pattern is a question. If the cell
contains a blank we replace it in place; otherwise the answer is written
to the cell immediately to the right. openpyxl round-trips the workbook,
so styles, column widths, merged cells, etc. are preserved.
"""
from __future__ import annotations

import json
import re
from typing import Dict, List

import openpyxl

from ..models import Anchor, Answer, DocFormat, EditResult, Question, QuestionType

_QUESTION_RE = re.compile(r"^\s*(\d+)\s*[.、)）]\s*(\S.*)$")
_BLANK_RE = re.compile(r"[_＿]{2,}")


def parse_questions(path: str) -> List[Question]:
    wb = openpyxl.load_workbook(path)
    questions: List[Question] = []
    counter = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                m = _QUESTION_RE.match(cell.value)
                if not m:
                    continue
                counter += 1
                number = m.group(1)
                blank = _BLANK_RE.search(cell.value)
                mode = "replace_blank" if blank else "right_cell"
                qtype = (
                    QuestionType.fill_blank if blank else QuestionType.short_answer
                )
                locator = json.dumps(
                    {
                        "sheet": ws.title,
                        "row": cell.row,
                        "col": cell.column,
                        "mode": mode,
                    }
                )
                questions.append(
                    Question(
                        id=f"q{number}",
                        number=number,
                        text=cell.value.strip(),
                        qtype=qtype,
                        anchor=Anchor(
                            format=DocFormat.xlsx, locator=locator,
                            detail={"mode": mode},
                        ),
                    )
                )
    wb.close()
    return questions


def write_answers(
    path: str, answers: List[Answer], questions: List[Question], output: str = None
) -> EditResult:
    wb = openpyxl.load_workbook(path)
    by_id: Dict[str, Answer] = {a.question_id: a for a in answers}
    written = 0
    for q in questions:
        ans = by_id.get(q.id)
        if ans is None:
            continue
        loc = json.loads(q.anchor.locator)
        ws = wb[loc["sheet"]]
        cell = ws.cell(row=loc["row"], column=loc["col"])
        if loc["mode"] == "replace_blank" and isinstance(cell.value, str) and _BLANK_RE.search(cell.value):
            cell.value = _BLANK_RE.sub(ans.text, cell.value, count=1)
        else:
            ws.cell(row=loc["row"], column=loc["col"] + 1).value = ans.text
        written += 1
    target = output or path
    wb.save(target)
    wb.close()
    return EditResult(
        path=target, format=DocFormat.xlsx, answers_written=written,
        in_place=(target == path),
    )
